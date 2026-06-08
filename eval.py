import os
import os.path as osp
import argparse
import openpyxl
import torch

from mmengine.runner import Runner
from mmengine.config import Config, DictAction

import segearthov3_segmentor
import segearthov3_change_detector
import custom_datasets
import custom_transforms


# force non-notebook backend
os.environ["MPLBACKEND"] = "Agg"


def parse_args():

    parser = argparse.ArgumentParser(
        description='CorrCLIP evaluation with MMSeg'
    )

    parser.add_argument(
        'config',
        nargs='?',
        default='./configs/cfg_potsdam.py'
    )

    parser.add_argument(
        '--show',
        action='store_true',
        help='show prediction results'
    )

    parser.add_argument(
        '--show_dir',
        default='./show_dir/',
        help='directory to save visualization images'
    )

    parser.add_argument(
        '--out',
        type=str,
        help='directory to save prediction output'
    )

    parser.add_argument(
        '--cfg-options',
        nargs='+',
        action=DictAction
    )

    parser.add_argument(
        '--launcher',
        choices=[
            'none',
            'pytorch',
            'slurm',
            'mpi'
        ],
        default='none'
    )

    parser.add_argument(
        '--local_rank',
        '--local-rank',
        type=int,
        default=0
    )

    args = parser.parse_args()

    if 'LOCAL_RANK' not in os.environ:
        os.environ[
            'LOCAL_RANK'
        ] = str(
            args.local_rank
        )

    return args


def append_experiment_result(
    file_path,
    experiment_data
):

    try:

        workbook = (
            openpyxl
            .load_workbook(
                file_path
            )
        )

    except FileNotFoundError:

        workbook = (
            openpyxl
            .Workbook()
        )

    sheet = workbook.active

    if sheet['A1'].value is None:

        sheet['A1'] = 'Model'
        sheet['B1'] = 'Dataset'
        sheet['C1'] = 'aAcc'
        sheet['D1'] = 'mIoU'
        sheet['E1'] = 'mAcc'

    last_row = (
        sheet.max_row
    )

    for index, result in enumerate(
        experiment_data,
        start=1
    ):

        sheet.cell(
            row=last_row + index,
            column=1,
            value=result['Model']
        )

        sheet.cell(
            row=last_row + index,
            column=2,
            value=result['Dataset']
        )

        sheet.cell(
            row=last_row + index,
            column=3,
            value=result['aAcc']
        )

        sheet.cell(
            row=last_row + index,
            column=4,
            value=result['mIoU']
        )

        sheet.cell(
            row=last_row + index,
            column=5,
            value=result['mAcc']
        )

    workbook.save(
        file_path
    )


def main():

    args = parse_args()

    print(
        os.getcwd()
    )

    cfg = Config.fromfile(
        args.config
    )

    cfg.launcher = (
        args.launcher
    )

    # output directory
    if args.out is not None:

        cfg.test_evaluator[
            'output_dir'
        ] = args.out

        cfg.test_evaluator[
            'keep_results'
        ] = True

    # cfg override
    if args.cfg_options is not None:

        cfg.merge_from_dict(
            args.cfg_options
        )

    cfg.work_dir = osp.join(
        './work_dirs',
        osp.splitext(
            osp.basename(
                args.config
            )
        )[0]
    )

    # -----------------------------------
    # create runner
    # -----------------------------------

    runner = Runner.from_cfg(
        cfg
    )

    # -----------------------------------
    # split dataset across GPUs
    # -----------------------------------

    gpu_count = int(
    os.environ.get(
        "TOTAL_GPUS",
        1
    )
)

    gpu_id = int(
        os.environ.get(
            "LOCAL_GPU_ID",
            0
        )
    )

    dataset = (
        runner
        .test_dataloader
        .dataset
    )

    all_data = (
        dataset.data_list
    )

    if gpu_count > 1:

        chunk_size = (
            len(all_data)
            + gpu_count
            - 1
        ) // gpu_count

        start = (
            gpu_id
            * chunk_size
        )

        end = min(
            start
            + chunk_size,
            len(all_data)
        )

        dataset.data_list = (
            all_data[
                start:end
            ]
        )

    print(
        f"GPU {gpu_id} "
        f"evaluating "
        f"{len(dataset.data_list)} "
        f"images"
    )

    for item in (
        dataset.data_list
    ):

        print(
            " -",
            osp.basename(
                item['img_path']
            )
        )

    # -----------------------------------
    # evaluation
    # -----------------------------------

    results = runner.test()

    results.update({

        'Model':
        cfg.model.model_type,

        'Dataset':
        cfg.dataset_type,

        'GPU':
        gpu_id
    })

    # -----------------------------------
    # save GPU-specific results
    # -----------------------------------

    result_file = (
        f"results_gpu"
        f"{gpu_id}.xlsx"
    )

    txt_file = (
        osp.join(
            cfg.work_dir,
            f"results_gpu"
            f"{gpu_id}.txt"
        )
    )

    append_experiment_result(
        result_file,
        [results]
    )

    with open(
        txt_file,
        'a'
    ) as f:

        f.write(
            osp.basename(
                args.config
            ).split('.')[0]
            + '\n'
        )

        for k, v in (
            results.items()
        ):

            f.write(
                k
                + ': '
                + str(v)
                + '\n'
            )


if __name__ == '__main__':
    main()